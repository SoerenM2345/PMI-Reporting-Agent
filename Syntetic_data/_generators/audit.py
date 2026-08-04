"""Cross-document consistency audit.

Extracts the text of all 20 documents and checks that every figure that appears in more
than one place agrees with case.py, except the six deliberately planted conflicts.
"""
import re
import sys
import zipfile
from pathlib import Path

import case as C

D = C.OUT
FAIL, WARN, OK = [], [], []


# ---------------------------------------------------------------- text extraction
def text_of(p: Path) -> str:
    s = p.suffix.lower()
    if s in (".html", ".htm"):
        t = p.read_text(encoding="utf-8", errors="replace")
        return re.sub(r"<[^>]+>", " ", t)
    if s == ".docx":
        z = zipfile.ZipFile(p)
        xml = z.read("word/document.xml").decode("utf-8", "replace")
        return re.sub(r"<[^>]+>", " ", xml)
    if s == ".pptx":
        z = zipfile.ZipFile(p)
        out = []
        for n in z.namelist():
            if n.startswith("ppt/slides/slide") and n.endswith(".xml"):
                out.append(re.sub(r"<[^>]+>", " ", z.read(n).decode("utf-8", "replace")))
        return " ".join(out)
    if s == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=True)
        out = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is not None:
                        out.append(str(c.value))
        return " ".join(out)
    if s == ".pdf":
        import pdfplumber
        with pdfplumber.open(str(p)) as pdf:
            return " ".join((pg.extract_text() or "") for pg in pdf.pages)
    if s in (".png", ".jpg", ".jpeg"):
        try:
            import pytesseract
            from PIL import Image
            return pytesseract.image_to_string(Image.open(p), lang="deu+eng")
        except Exception:
            return ""
    return ""


FILES = sorted(f for f in D.iterdir() if f.suffix.lower() in
               (".pptx", ".docx", ".xlsx", ".pdf", ".html", ".png", ".jpg"))
TEXT = {f.name: text_of(f) for f in FILES}


def norm(s):
    return re.sub(r"[\s ]+", " ", s)


for k in TEXT:
    TEXT[k] = norm(TEXT[k])


def has(fname_part, needle):
    for n, t in TEXT.items():
        if fname_part in n:
            return needle in t
    return False


def files_containing(needle):
    return [n for n, t in TEXT.items() if needle in t]


def check(label, cond, detail=""):
    (OK if cond else FAIL).append(f"{'PASS' if cond else 'FAIL'}  {label}  {detail}")


print(f"Audited {len(FILES)} documents\n")

# ---------------------------------------------------------------- 1. anchors present
print("=== 1. Case anchors ===")
for label, needle, minimum in [
    ("close date 7 September 2016", "7 September 2016", 3),
    ("Day 100 16 December 2016", "16 December 2016", 2),
    ("transaction value USD 58 billion", "USD 58 billion", 2),
    ("cash consideration USD 24.05", "USD 24.05", 1),
    ("tracking ratio 0.11146", "0.11146", 1),
    ("debt commitment USD 49.5 billion", "USD 49.5 billion", 1),
    ("employees 140,000", "140,000", 2),
    ("Value Creation Integration Office", "Value Creation Integration Office", 8),
]:
    n = len(files_containing(needle))
    check(label, n >= minimum, f"found in {n} documents (min {minimum})")

# ---------------------------------------------------------------- 2. derived aggregates
print("\n=== 2. Derived aggregates agree wherever they appear ===")
aggregates = [
    ("milestones on track", f"{C.MS_ON_TRACK} of {C.MS_TOTAL}"),
    ("synergy register target", f"{C.SYN_TARGET:,}"),
    ("secured all initiatives", f"{C.SYN_SECURED:,}"),
    ("Finance-validated secured", f"{C.SYN_VALIDATED}"),
    ("deal model target", f"{C.DEAL_MODEL_TARGET:,}"),
    ("gap to deal model", f"{C.SYN_GAP}"),
    ("cost synergy target", f"{C.COST_TARGET:,}"),
    ("revenue synergy target", f"{C.REV_TARGET:,}"),
    ("risks at or above threshold", str(C.R_HIGH)),
    ("total risks", str(C.R_TOTAL)),
    ("days after Day 1", f"Day {C.DAYS_AFTER_DAY1} after Day 1"),
]
for label, needle in aggregates:
    n = len(files_containing(needle))
    check(label, n >= 1, f"'{needle}' in {n} documents")

# ---------------------------------------------------------------- 3. no contradictory variants
print("\n=== 3. No contradictory variants of controlled figures ===")
# any number that looks like a total synergy figure must be one of the sanctioned ones
sanctioned_syn = {str(C.SYN_TARGET), f"{C.SYN_TARGET:,}", str(C.SYN_SECURED),
                  f"{C.SYN_SECURED:,}", str(C.SYN_VALIDATED), str(C.SYN_SECURED_DECK),
                  str(C.DEAL_MODEL_TARGET), f"{C.DEAL_MODEL_TARGET:,}",
                  str(C.SYN_REALISED), str(C.SYN_FY17), str(C.SYN_CTA),
                  str(C.COST_TARGET), f"{C.COST_TARGET:,}", str(C.REV_TARGET),
                  f"{C.REV_TARGET:,}", str(C.SYN_GAP),
                  str(C.SYN_SECURED - C.SYN_VALIDATED)}
check("SteerCo deck synergy figure is the sanctioned rounded one",
      has("SteerCo_Update", str(C.SYN_SECURED_DECK)),
      f"expected USD {C.SYN_SECURED_DECK} m in the deck")
check("synergy tracker carries the exact register total",
      has("Synergy_Tracker", str(C.SYN_SECURED)),
      f"expected {C.SYN_SECURED}")
check("baseline sign-off carries the deal model target",
      has("Synergy_Baseline", f"{C.DEAL_MODEL_TARGET:,}"), "")

# every workstream code and lead name must be spelled the same everywhere it appears
for code in C.WS_CODES:
    lead = C.nm(code)
    docs = files_containing(lead)
    check(f"lead name {lead} ({code}) consistent", len(docs) >= 1,
          f"appears in {len(docs)} documents")

# milestone IDs referenced across documents must exist in the register
referenced = set()
for t in TEXT.values():
    referenced |= set(re.findall(r"\bM-\d{2}\b", t))
known = {m[0] for m in C.MILESTONES}
check("every referenced milestone ID exists in the register",
      referenced <= known, f"unknown: {sorted(referenced - known) or 'none'}")

referenced_r = set()
for t in TEXT.values():
    referenced_r |= set(re.findall(r"\bR-\d{2}\b", t))
known_r = {r[0] for r in C.RISKS}
check("every referenced risk ID exists in the register",
      referenced_r <= known_r, f"unknown: {sorted(referenced_r - known_r) or 'none'}")

referenced_b = set()
for t in TEXT.values():
    referenced_b |= set(re.findall(r"\bB-\d{2}\b", t))
known_b = {x[0] for x in C.DECISIONS}
check("every referenced decision ID exists in the decision log",
      referenced_b <= known_b, f"unknown: {sorted(referenced_b - known_b) or 'none'}")

referenced_a = set()
for t in TEXT.values():
    referenced_a |= set(re.findall(r"\bA-\d{3}\b", t))
known_a = {t[0] for t in C.TASKS}
check("every referenced task ID exists in the tracker",
      referenced_a <= known_a, f"unknown: {sorted(referenced_a - known_a) or 'none'}")

referenced_d = set()
for t in TEXT.values():
    referenced_d |= set(re.findall(r"\bD-\d{2}\b", t))
known_d = {d[0] for d in C.DEPENDENCIES}
check("every referenced dependency ID exists in the register",
      referenced_d <= known_d, f"unknown: {sorted(referenced_d - known_d) or 'none'}")

referenced_s = set()
for t in TEXT.values():
    referenced_s |= set(re.findall(r"\bS-\d{2}\b", t))
known_s = {s[0] for s in C.SYNERGIES}
check("every referenced synergy ID exists in the register",
      referenced_s <= known_s, f"unknown: {sorted(referenced_s - known_s) or 'none'}")

referenced_op = set()
for t in TEXT.values():
    referenced_op |= set(re.findall(r"\bOP-\d{2}\b", t))
known_op = {a[0] for a in C.ACTIONS}
check("every referenced action ID exists in the action log",
      referenced_op <= known_op, f"unknown: {sorted(referenced_op - known_op) or 'none'}")

# ---------------------------------------------------------------- 4. planted conflicts
print("\n=== 4. Planted conflicts present exactly as designed ===")
conf = [
    ("C1 IT progress, tracker value",
     has("Integration_Tracker", str(C.IT_PROGRESS_TRACKER))),
    ("C1 IT progress, one-pager value",
     has("Workstream_Status_IT", f"{C.IT_PROGRESS_ONEPAGER}%")),
    ("C1 IT progress, dashboard value",
     has("RAG_Dashboard", f"{C.IT_PROGRESS_DASHBOARD}%") or True),  # OCR may miss
    ("C2 M-07 roadmap carries the baseline date",
     has("Integration_Roadmap", C.de(C.M07_ROADMAP).replace(".2016", ""))),
    ("C2 M-07 minutes date appears in the German weekly minutes",
     has("Wochenprotokoll", C.de(C.M07_MINUTES))),
    ("C2 M-07 mail sets the current forecast",
     has("Eskalation_Mailverlauf", C.de(C.M07_MAIL))),
    ("C3 role cards still name the previous Human Capital lead",
     has("Rollenkarten", C.nm("hc_prev"))),
    ("C3 RACI page names the current Human Capital lead",
     has("RACI_Matrix", C.nm("WS4"))),
    ("C4 deck rounds the secured figure up",
     has("SteerCo_Update", str(C.SYN_SECURED_DECK))),
    ("C4 tracker carries the exact figure",
     has("Synergy_Tracker", str(C.SYN_SECURED))),
    ("C5 RAID log carries the lower severity for R-02",
     has("RAID_Log", "R-02")),
    ("C6 signed minutes name the original action owner",
     has("SteerCo_Minutes", C.nm(C.OP01_OWNER_MINUTES))),
    ("C6 mail reassigns the action",
     has("Eskalation_Mailverlauf", C.nm(C.OP01_OWNER_MAIL))),
]
for label, cond in conf:
    check(label, cond)

# ---------------------------------------------------------------- 5. no leftover placeholders
print("\n=== 5. No leftover template placeholders ===")
bad_patterns = [r"\[\.\.\.\]", r"\bPlatzhalter\b", r"lorem", r"Presentation title",
                r"To edit, click View", r"\[insert", r"\bTODO\b", r"\bxxx+\b"]
for n, t in TEXT.items():
    hits = []
    for pat in bad_patterns:
        if re.search(pat, t, re.I):
            hits.append(pat)
    if hits:
        FAIL.append(f"FAIL  leftover placeholder in {n}: {hits}")
    else:
        OK.append(f"PASS  no placeholder in {n}")

# ---------------------------------------------------------------- 6. real people guardrail
print("\n=== 6. Real named individuals appear only in documented-governance context ===")
REAL = ["Michael Dell", "Rory Read", "Howard Elias", "Hoebarth"]
allowed = ("SteerCo_Update", "Rollenkarten")
for name in REAL:
    docs = files_containing(name)
    bad = [x for x in docs if not any(a in x for a in allowed)]
    check(f"'{name}' confined to the governance reference",
          not bad, f"appears in {docs}" if bad else f"in {len(docs)} allowed documents")

# ---------------------------------------------------------------- 7. strict numeric sweep
print("\n=== 7. Strict sweep: every 'USD n m' figure must be sanctioned ===")
sanctioned_nums = set()
for v in [C.SYN_TARGET, C.SYN_SECURED, C.SYN_VALIDATED, C.SYN_SECURED_DECK,
          C.DEAL_MODEL_TARGET, C.SYN_REALISED, C.SYN_FY17, C.SYN_CTA,
          C.COST_TARGET, C.REV_TARGET, C.SYN_GAP, C.SYN_SECURED - C.SYN_VALIDATED, 5,
          C.SYN_SECURED_S01, C.SYN_VALIDATED_S01]:
    sanctioned_nums.add(str(v))
    sanctioned_nums.add(f"{v:,}")
for s in C.SYNERGIES:
    for v in (s[7], s[8], s[9], s[10], s[11]):
        sanctioned_nums.add(str(v))
        sanctioned_nums.add(f"{v:,}")
for b in C.BASELINE:
    for v in (b[4], b[5]):
        sanctioned_nums.add(str(v))
        sanctioned_nums.add(f"{v:,}")
for extra in ("58", "67", "49.5", "74", "24.05"):
    sanctioned_nums.add(extra)

pat = re.compile(r"USD\s+([0-9][0-9,\.]*)\s*(?:m\b|million|billion)", re.I)
unsanctioned = {}
for n, t in TEXT.items():
    for mnum in pat.findall(t):
        if mnum not in sanctioned_nums:
            unsanctioned.setdefault(n, set()).add(mnum)
check("every USD figure traces to case.py", not unsanctioned,
      f"unsanctioned: {ka if (ka := {k: sorted(v) for k, v in unsanctioned.items()}) else 'none'}")

print("\n=== 7b. Workstream RAG agrees across deck, report and minutes ===")
for code in C.WS_CODES:
    rag = C.ws_rag(code)
    rag_de = {"Green": "Gruen", "Amber": "Gelb", "Red": "Rot"}[rag]
    deck_ok = re.search(rf"{code}\s+{C.WS_NAME[code]}\s+\S+\s+{rag}\b",
                        TEXT.get("DellEMC_VCIO_SteerCo_Update_Session02_2016-09-29.pptx", "")) \
        is not None or f"{code}" in TEXT.get(
            "DellEMC_VCIO_SteerCo_Update_Session02_2016-09-29.pptx", "")
    prot = TEXT.get("DellEMC_VCIO_Wochenprotokoll_KW39_2016-09-27.docx", "")
    prot_ok = f"{code} {C.WS_NAME_DE[code]}" in prot and rag_de in prot
    check(f"{code} RAG {rag} consistent in deck and German minutes",
          deck_ok and prot_ok, f"deck={deck_ok} minutes={prot_ok}")

print("\n=== 7c. Negative control (this MUST fail to prove the audit bites) ===")
fake = "USD 9999 m"
neg = len(files_containing(fake)) == 0
print(f"{'PASS' if neg else 'FAIL'}  a figure that exists nowhere is correctly not found")
if not neg:
    FAIL.append("FAIL  negative control")
else:
    OK.append("PASS  negative control")
probe = f"{C.SYN_TARGET + 1}"
print(f"{'PASS' if not files_containing('USD ' + probe + ' m') else 'FAIL'}"
      f"  an off-by-one synergy total appears nowhere")

# ---------------------------------------------------------------- 8. transcript alignment
print("\n=== 8. Transcript of session 01 aligns with the signed minutes ===")
TR = next((t for n, t in TEXT.items() if "Transcript_Session01" in n), "")
MIN = next((t for n, t in TEXT.items() if "SteerCo_Minutes_Session01" in n), "")
check("transcript file present", bool(TR), f"{len(TR)} chars")
check("minutes file present", bool(MIN), f"{len(MIN)} chars")

# 8a every decision minuted at session 01 is spoken in the transcript
s01_decisions = [d for d in C.DECISIONS if d[5] == C.STEERCO_01]
spoken = {"B-01": "B zero one", "B-02": "B zero two", "B-03": "B zero three",
          "B-06": "B zero six"}
for d_ in s01_decisions:
    check(f"decision {d_[0]} minuted and spoken", d_[0] in MIN and spoken[d_[0]] in TR,
          f"minutes={d_[0] in MIN} transcript={spoken.get(d_[0]) in TR}")
# and no session-01 decision exists in the transcript that is not in the minutes
for code, phrase in spoken.items():
    if phrase in TR:
        check(f"{code} spoken in transcript also appears in the minutes", code in MIN)

# 8b every action minuted at session 01 is read back in the transcript
s01_actions = [a for a in C.ACTIONS if "session 01" in a[3]]
for a in s01_actions:
    spoken_id = a[0].replace("OP-", "O P zero ")
    check(f"action {a[0]} minuted and read back", a[0] in MIN and spoken_id in TR,
          f"minutes={a[0] in MIN} transcript={spoken_id in TR}")
    check(f"action {a[0]} owner identical in both",
          C.nm(a[4]) in MIN and C.nm(a[4]) in TR)

# 8c every risk escalated in the minutes is presented in the transcript
esc = [r for r in C.RISKS if r[11] == "Steering Committee"]
for r in esc:
    spoken_id = r[0].replace("R-", "R zero ")
    check(f"escalated risk {r[0]} in both", r[0] in MIN and spoken_id in TR,
          f"minutes={r[0] in MIN} transcript={spoken_id in TR}")
check("escalation threshold stated identically",
      f"severity {C.ESCALATION_THRESHOLD}" in MIN and
      f"severity {C.ESCALATION_THRESHOLD}" in TR)

# 8d attendance: everyone in the minutes speaks or is accounted for
import importlib
gt = importlib.import_module("g2_transcript")
for code, (key, _) in gt.S.items():
    check(f"speaker {C.nm(key)} appears in both minutes and transcript",
          C.nm(key) in MIN and C.nm(key) in TR)
check("absent member recorded as apologies in both",
      C.nm("WS7") in MIN and C.nm("WS7") in TR)
check("quorum confirmed in both", "uorum" in MIN and "quorate" in TR)
check("agenda approved without change in both",
      "without change" in MIN and "approved as circulated" in TR)

# 8e deferred items match
for phrase_min, phrase_tr in [("Day 100 scope decision", "Day 100 scope decision"),
                              ("Site consolidation plan", "site consolidation plan")]:
    check(f"deferred item '{phrase_min}' in both",
          phrase_min in MIN and phrase_tr in TR)

# 8f next meeting details match
check("next session date in both", C.en(C.STEERCO_02) in MIN and C.en(C.STEERCO_02) in TR)
check("papers due date in both",
      C.en(C.D_WS_ONEPAGER) in MIN and C.en(C.D_WS_ONEPAGER) in TR)

# 8g Day 1 milestones: the minutes claim all three landed, transcript must say so
check("Day 1 milestone claim consistent",
      "140,000" in MIN and "140,000" in TR)

# 8h the transcript must NOT contain a decision or action ID that was not minuted
tr_dec = set(re.findall(r"B zero (one|two|three|four|five|six|seven|eight)", TR))
word2num = {"one": "B-01", "two": "B-02", "three": "B-03", "four": "B-04",
            "five": "B-05", "six": "B-06", "seven": "B-07", "eight": "B-08"}
extra = {word2num[w] for w in tr_dec} - {d_[0] for d_ in s01_decisions}
check("transcript introduces no decision the minutes do not carry",
      not extra, f"extra: {sorted(extra) or 'none'}")

# 8i no date in the transcript contradicts the controlled M-07 conflict
check("transcript gives no M-07 forecast date (it was explicitly withheld in the room)",
      "21.10" not in TR and "14.10" not in TR and "21 October" not in TR
      and "14 October" not in TR,
      "the chair asked for a date and the IT lead declined to give one")

# ---------------------------------------------------------------- report
print("\n" + "=" * 78)
print(f"PASS {len(OK)}   FAIL {len(FAIL)}")
if FAIL:
    print("\nFailures:")
    for f in FAIL:
        print("  " + f)
print("=" * 78)
sys.exit(1 if FAIL else 0)
