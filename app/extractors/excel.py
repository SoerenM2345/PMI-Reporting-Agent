"""Excel extractor (spec §5.1). The highest-priority source: a tracker is the
system of record, so its values win most conflicts (§9).

Two things the original version got wrong, both of which silently lost data:

* **One table per sheet.** PMI trackers routinely stack two or three tables on one
  sheet ("Open Risks" above "Closed Risks"), separated by a blank row. Reading only
  the first header row dropped everything below the second one.
* **Formulas.** pandas reads the cached value, which is right — but when a workbook
  has never been opened by Excel there *is* no cached value, and the cell reads as
  the formula string. We detect that and say so rather than treating "=SUM(B2:B9)"
  as a project figure.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator

import pandas as pd

from app.extractors.base import (
    classify_table,
    find_progress_mentions,
    make_source,
    normalize_header,
    rows_to_records,
)
from app.models.pmi import ExtractionMethod, SourceFormat

suffixes: tuple[str, ...] = (".xlsx", ".xlsm", ".xls")
format: SourceFormat = SourceFormat.EXCEL

#: A header row needs at least this many recognisable columns.
_MIN_HEADER_HITS = 2
#: Blank rows separating stacked tables.
_GAP = 1


def extract(path: Path) -> list[dict]:
    records: list[dict] = []
    workbook = pd.ExcelFile(path)

    for sheet in workbook.sheet_names:
        frame = workbook.parse(sheet, header=None)
        if frame.empty:
            continue

        for header_idx, end_idx in _find_tables(frame):
            headers = [
                str(c) if pd.notna(c) else "" for c in frame.iloc[header_idx]
            ]
            body = frame.iloc[header_idx + 1: end_idx]
            rows = [
                [None if pd.isna(c) else c for c in row]
                for row in body.itertuples(index=False)
            ]
            if not rows:
                continue

            record_type = classify_table(headers, context=sheet)
            source = make_source(
                path.name, format,
                sheet_name=sheet,
                extraction_method=ExtractionMethod.TABLE_PARSE,
            )
            # +2: spreadsheet rows are 1-based and the header occupies one of them.
            records.extend(rows_to_records(
                headers, rows, record_type, source,
                first_data_row=header_idx + 2,
            ))

        # Stray summary cells: "Overall Progress: 82%" written outside any table.
        # This is how the spec's own 82-vs-75 example enters the model.
        flat = _flatten(frame)
        if flat:
            source = make_source(
                path.name, format,
                sheet_name=sheet,
                extraction_method=ExtractionMethod.TEXT_REGEX,
            )
            for pct in find_progress_mentions(flat):
                records.append({
                    "type": "kpi", "name": "Overall Progress", "value": pct,
                    "unit": "%", "source": source,
                })

    records.extend(_formula_warnings(path))
    return records


# ------------------------------------------------------------------- internals
def _find_tables(frame: pd.DataFrame) -> Iterator[tuple[int, int]]:
    """Yield (header_row, end_row) for every table stacked on a sheet.

    A table starts at a row with >= 2 recognisable headers and runs until a blank
    row or the next header row.
    """
    header_rows = [
        i for i in range(len(frame)) if _looks_like_a_header(frame.iloc[i])
    ]
    if not header_rows:
        return

    for position, header_idx in enumerate(header_rows):
        # Stop at the next header, or at the first run of blank rows.
        limit = (
            header_rows[position + 1]
            if position + 1 < len(header_rows)
            else len(frame)
        )
        end = limit
        blanks = 0
        for i in range(header_idx + 1, limit):
            if frame.iloc[i].isna().all():
                blanks += 1
                if blanks > _GAP:
                    end = i - blanks + 1
                    break
            else:
                blanks = 0

        if end > header_idx + 1:
            yield header_idx, end


def _looks_like_a_header(row) -> bool:
    """A header row is text. A row containing a date or a number is data.

    Without this, a task row whose status happens to resemble a column name gets
    promoted to a header and the table is split at that row — losing every row below
    it, silently. Requiring the row to be *entirely* non-numeric is what makes header
    detection safe enough to run on every row of the sheet.
    """
    hits = 0
    for cell in row:
        if pd.isna(cell):
            continue
        if isinstance(cell, (int, float, pd.Timestamp)) and not isinstance(cell, bool):
            return False  # numbers and dates never appear in a header row
        if normalize_header(cell):
            hits += 1
    return hits >= _MIN_HEADER_HITS


def _flatten(frame: pd.DataFrame) -> str:
    return " ".join(str(c) for c in frame.to_numpy().flatten() if pd.notna(c))


def _formula_warnings(path: Path) -> list[dict]:
    """Report formulas with no cached result (§21.17 — say what we could not read).

    openpyxl in data_only mode returns None for a formula cell that Excel has never
    evaluated. Treating that as "no value" is right; treating the formula text as a
    number would be a fabrication.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - openpyxl is a declared dependency
        return []

    if path.suffix.lower() == ".xls":  # legacy format, openpyxl cannot read it
        return []

    try:
        cached = load_workbook(path, data_only=True, read_only=True)
        raw = load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return []

    unevaluated: list[str] = []
    try:
        for sheet_name in raw.sheetnames:
            raw_sheet, cached_sheet = raw[sheet_name], cached[sheet_name]
            for row in raw_sheet.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or not cell.value.startswith("="):
                        continue
                    if cached_sheet[cell.coordinate].value is None:
                        unevaluated.append(f"{sheet_name}!{cell.coordinate}")
    except Exception:
        return []
    finally:
        cached.close()
        raw.close()

    if not unevaluated:
        return []

    shown = ", ".join(unevaluated[:5])
    more = f" (and {len(unevaluated) - 5} more)" if len(unevaluated) > 5 else ""
    return [{
        "type": "note",
        "is_warning": True,
        "text": (
            f"{len(unevaluated)} formula cell(s) have no cached value and were read "
            f"as empty: {shown}{more}. Open and save the workbook in Excel to "
            f"populate them."
        ),
        "source": make_source(path.name, format,
                              extraction_method=ExtractionMethod.DERIVED),
    }]
