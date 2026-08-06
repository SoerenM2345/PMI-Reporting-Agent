"""Populated XLSX: integration tracker (DE headers), RAID log (EN), synergy tracker (EN)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import case as C

A = "Arial"
HDR = PatternFill("solid", fgColor=C.D_DARK)
CALC = PatternFill("solid", fgColor=C.D_PALE)
ALT = PatternFill("solid", fgColor="F7F9F3")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RED = Font(A, 9, bold=True, color=C.RAG_COLOR["Red"])
AMB = Font(A, 9, bold=True, color=C.RAG_COLOR["Amber"])

STATUS_DE = {"Open": "Offen", "In progress": "In Arbeit", "Done": "Erledigt",
             "Blocked": "Blockiert", "Overdue": "Überfällig"}
PRIO_DE = {"High": "Hoch", "Medium": "Mittel", "Low": "Niedrig"}
MS_DE = {"done": "Erledigt", "on_track": "Im Plan", "at_risk": "Gefährdet",
         "delayed": "Verzögert"}


def sheet(wb, name, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = name
    ws.sheet_view.showGridLines = False
    return ws


def titleblock(ws, title, sub):
    ws["A1"] = title
    ws["A1"].font = Font(A, 13, bold=True, color=C.D_DARK)
    ws["A2"] = sub
    ws["A2"].font = Font(A, 8.5, color=C.D_GREY)


def header(ws, row, heads, widths):
    for j, h in enumerate(heads, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(A, 8.5, bold=True, color="FFFFFF")
        c.fill = HDR
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BOX
        ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def put(ws, row, values, red_at=None, calc_cols=(), date_cols=()):
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(A, 9)
        c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if row % 2 == 0:
            c.fill = ALT
        if j in calc_cols:
            c.fill = CALC
        if j in date_cols:
            c.number_format = "DD.MM.YYYY"
        if red_at and j == red_at[0]:
            if str(v) in ("Überfällig", "Overdue", "Blockiert", "Blocked", "Verzögert",
                          "Delayed", "Red", "Rot", "Critical", "High"):
                c.font = RED
            elif str(v) in ("Gefährdet", "At risk", "Amber", "Gelb", "Medium"):
                c.font = AMB


def legend(ws, lines, notes):
    titleblock(ws, f"{C.PROGRAM} - {C.OFFICE} ({C.OFFICE_ABBR})",
               f"{C.ACQUIRER} and {C.TARGET}  |  Reporting week {C.WEEK_LABEL}  |  "
               f"Day {C.DAYS_AFTER_DAY1} after Day 1  |  Close {C.en(C.DAY1)}")
    r = 4
    ws.cell(row=r, column=1, value="How to use this workbook").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    r += 1
    for line in lines:
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(A, 9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Conventions and assumptions").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    r += 1
    for line in notes:
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(A, 9, color=C.D_GREY)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Sources for the case anchoring").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    r += 1
    for s in C.SOURCES:
        ws.cell(row=r, column=1, value=s).font = Font(A, 8.5, color=C.D_GREY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
    r += 1
    c = ws.cell(row=r, column=1,
                value="Synthetic workbook. All operational names, figures, risks and synergy "
                      "values are invented and internally consistent; they are not a "
                      "representation of what any party actually did.")
    c.font = Font(A, 8.5, italic=True, color=C.D_GREY)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.column_dimensions["A"].width = 34
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 15


# =====================================================================
# 1. Integration Tracker (German headers)
# =====================================================================
def tracker():
    wb = Workbook()
    ws = sheet(wb, "Legende", first=True)
    legend(ws,
           ["1. Register Massnahmenplan: je Zeile eine Aufgabe. Fuehrende Quelle fuer "
            "Aufgaben und Termine nach Beschluss B-02.",
            "2. Register Meilensteine: Basistermin bleibt nach Freigabe unveraendert. "
            "Verschiebungen laufen ueber den Prognosetermin, die Abweichung rechnet sich.",
            "3. Register Beschlussregister und Massnahmenverfolgung: Beschluesse und offene "
            "Punkte aus VCIO Weekly und Steering Committee.",
            "4. Register Abhaengigkeiten: jede Uebergabe zwischen Teilprojekten muss nach "
            "Beschluss B-03 von beiden Leitungen bestaetigt sein.",
            "5. Register Auswertung rechnet automatisch. Formeln bitte nicht ueberschreiben."],
           [f"Berichtswoche {C.WEEK_LABEL}, Tag {C.DAYS_AFTER_DAY1} nach Day 1 "
            f"({C.de(C.DAY1)}). Day 100 ist der {C.de(C.DAY100)}.",
            "Datumsformat durchgaengig TT.MM.JJJJ. Fortschritt in Prozent als ganze Zahl.",
            f"Sieben Teilprojekte berichten an das Steering Committee. Das Programm fuehrt "
            f"insgesamt {C.WORKSTREAM_COUNT_TOTAL} Teilprojekte.",
            "Bei Abweichungen zwischen dieser Datei und einer Statusfolie gilt diese Datei."])

    # Massnahmenplan
    ws = sheet(wb, "Massnahmenplan")
    titleblock(ws, "Massnahmenplan (Master Task List)",
               f"Stand {C.de(C.D_TRACKER)}  |  {len(C.TASKS)} Aufgaben  |  "
               f"Verantwortlich: {C.OFFICE} ({C.OFFICE_ABBR})")
    header(ws, 4, ["ID", "Aufgabe", "Teilprojekt", "Verantwortlich", "Faellig am", "Status",
                   "Fortschritt %", "Prioritaet", "Abhaengig von", "Kommentar"],
           [9, 52, 13, 18, 12, 13, 12, 11, 13, 30])
    for i, t in enumerate(C.TASKS):
        put(ws, 5 + i, [t[0], t[2], f"{t[3]} {C.WS_NAME_DE[t[3]]}", C.nm(t[4]), t[5],
                        STATUS_DE[t[6]], t[7], PRIO_DE[t[8]], t[9] or "",
                        t[1]], red_at=(6,), date_cols=(5,))

    # Meilensteine
    ws = sheet(wb, "Meilensteine")
    titleblock(ws, "Meilensteinplan (Milestone Plan)",
               f"Day 1 {C.de(C.DAY1)}  |  Day 30 {C.de(C.DAY30)}  |  Day 100 {C.de(C.DAY100)}"
               f"  |  {C.MS_GATE} von {C.MS_TOTAL} gate-relevant")
    header(ws, 4, ["ID", "Meilenstein", "Phase", "Teilprojekt", "Verantwortlich",
                   "Basistermin", "Prognosetermin", "Abweichung (Tage)", "Status",
                   "Gate-relevant", "Kommentar"],
           [8, 48, 17, 13, 18, 13, 15, 15, 13, 13, 46])
    for i, m in enumerate(C.MILESTONES):
        r = 5 + i
        put(ws, r, [m[0], m[2], m[5], f"{m[3]} {C.WS_NAME_DE[m[3]]}",
                    C.nm(m[4]) if m[4] in C.PEOPLE else "", m[6], m[7], None,
                    MS_DE[m[8]], "Ja" if m[9] else "Nein", m[10]],
            red_at=(9,), date_cols=(6, 7))
        c = ws.cell(row=r, column=8, value=f"=G{r}-F{r}")
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX

    # Beschlussregister
    ws = sheet(wb, "Beschlussregister")
    titleblock(ws, "Beschlussregister (Decision Log)",
               f"Beschluesse aus VCIO Weekly und Steering Committee, Stand {C.de(C.D_TRACKER)}")
    header(ws, 4, ["ID", "Beschluss", "Begruendung", "Gremium", "Datum",
                   "Betroffene Teilprojekte", "Umsetzung durch", "Umsetzungsstand"],
           [9, 52, 48, 18, 12, 22, 18, 16])
    for i, x in enumerate(C.DECISIONS):
        put(ws, 5 + i, [x[0], x[2], x[3], x[4], x[5], x[6], C.nm(x[7]), x[8]],
            date_cols=(5,))

    # Massnahmenverfolgung
    ws = sheet(wb, "Massnahmenverfolgung")
    titleblock(ws, "Massnahmenverfolgung (Action Item Log)",
               f"Offene Punkte aus Sitzungen, Stand {C.de(C.D_TRACKER)}  |  "
               f"{C.ACT_DONE} von {C.ACT_TOTAL} geschlossen, {C.ACT_SHIFTED} verschoben")
    header(ws, 4, ["Nr.", "Offener Punkt", "Quelle", "Verantwortlich", "Teilprojekt",
                   "Urspruenglich faellig", "Neuer Termin", "Verschiebungen", "Status"],
           [9, 52, 26, 18, 13, 17, 14, 15, 14])
    for i, a in enumerate(C.ACTIONS):
        put(ws, 5 + i, [a[0], a[2], a[3], C.nm(a[4]), a[5], a[6], a[7], a[8],
                        STATUS_DE[a[9]]], red_at=(9,), date_cols=(6, 7))

    # Abhaengigkeiten
    ws = sheet(wb, "Abhaengigkeiten")
    titleblock(ws, "Abhaengigkeitsregister (Dependency Register)",
               f"Uebergaben zwischen Teilprojekten  |  Beschluss B-03 verlangt Bestaetigung "
               f"durch beide Leitungen  |  {C.DEP_UNCONFIRMED_CRITICAL} offen")
    header(ws, 4, ["ID", "Uebergabe", "Von Teilprojekt", "An Teilprojekt", "Benoetigt bis",
                   "Kritikalitaet", "Von beiden bestaetigt", "Status"],
           [9, 58, 16, 16, 14, 14, 19, 14])
    for i, d in enumerate(C.DEPENDENCIES):
        put(ws, 5 + i, [d[0], d[1], f"{d[2]} {C.WS_NAME_DE[d[2]]}",
                        f"{d[3]} {C.WS_NAME_DE[d[3]]}", d[4], d[5],
                        "Ja" if d[6] == "Yes" else "Nein", d[7]],
            red_at=(8,), date_cols=(5,))

    # Auswertung
    ws = sheet(wb, "Auswertung")
    titleblock(ws, "Auswertung (Summary)",
               "Rechnet automatisch aus den Registern Massnahmenplan und Meilensteine. "
               "Keine Eingabe.")
    last_t = 4 + len(C.TASKS)
    last_m = 4 + len(C.MILESTONES)
    ws.cell(row=4, column=1, value="Aufgaben nach Status").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    header(ws, 5, ["Status", "Anzahl", "Anteil"], [26, 14, 12])
    stat = ["Offen", "In Arbeit", "Erledigt", "Blockiert", "Überfällig"]
    for i, s in enumerate(stat):
        r = 6 + i
        ws.cell(row=r, column=1, value=s).font = Font(A, 9)
        for col, f in ((2, f"=COUNTIF(Massnahmenplan!$F$5:$F${last_t},$A{r})"),
                       (3, f'=IFERROR($B{r}/$B$11,"")')):
            c = ws.cell(row=r, column=col, value=f)
            c.fill = CALC
            c.font = Font(A, 9)
            c.border = BOX
            if col == 3:
                c.number_format = "0%"
    ws.cell(row=11, column=1, value="Summe").font = Font(A, 9, bold=True)
    c = ws.cell(row=11, column=2, value="=SUM($B$6:$B$10)")
    c.fill = CALC
    c.font = Font(A, 9, bold=True)
    c.border = BOX

    ws.cell(row=13, column=1, value="Je Teilprojekt").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    header(ws, 14, ["Teilprojekt", "Aufgaben", "davon erledigt", "davon ueberfaellig",
                    "Fortschritt %", "Meilensteine", "davon verzoegert"],
           [30, 12, 15, 17, 14, 14, 17])
    for i, code in enumerate(C.WS_CODES):
        r = 15 + i
        label = f"{code} {C.WS_NAME_DE[code]}"
        ws.cell(row=r, column=1, value=label).font = Font(A, 9)
        for col, f in (
            (2, f'=COUNTIF(Massnahmenplan!$C$5:$C${last_t},$A{r})'),
            (3, f'=COUNTIFS(Massnahmenplan!$C$5:$C${last_t},$A{r},'
                f'Massnahmenplan!$F$5:$F${last_t},"Erledigt")'),
            (4, f'=COUNTIFS(Massnahmenplan!$C$5:$C${last_t},$A{r},'
                f'Massnahmenplan!$F$5:$F${last_t},"Überfällig")'),
            (5, f'=IFERROR(ROUND(AVERAGEIF(Massnahmenplan!$C$5:$C${last_t},$A{r},'
                f'Massnahmenplan!$G$5:$G${last_t}),0),"")'),
            (6, f'=COUNTIF(Meilensteine!$D$5:$D${last_m},$A{r})'),
            (7, f'=COUNTIFS(Meilensteine!$D$5:$D${last_m},$A{r},'
                f'Meilensteine!$I$5:$I${last_m},"Verzögert")'),
        ):
            c = ws.cell(row=r, column=col, value=f)
            c.fill = CALC
            c.font = Font(A, 9)
            c.border = BOX

    ws.cell(row=23, column=1, value="Meilensteine gesamt").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    for i, (lab, f) in enumerate([
        ("Meilensteine gesamt", f'=COUNTA(Meilensteine!$B$5:$B${last_m})'),
        ("davon gate-relevant", f'=COUNTIF(Meilensteine!$J$5:$J${last_m},"Ja")'),
        ("davon verzoegert", f'=COUNTIF(Meilensteine!$I$5:$I${last_m},"Verzögert")'),
        ("groesste Abweichung in Tagen", f'=MAX(Meilensteine!$H$5:$H${last_m})'),
        ("Tage bis Day 100", f'=DATE(2016,12,16)-DATE(2016,9,29)'),
    ]):
        r = 24 + i
        ws.cell(row=r, column=1, value=lab).font = Font(A, 9)
        c = ws.cell(row=r, column=2, value=f)
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX
    ws.column_dimensions["A"].width = 32

    wb.save(C.OUT / "DellEMC_VCIO_Integration_Tracker_W3_2016-09-29.xlsx")


# =====================================================================
# 2. RAID Log (English)
# =====================================================================
def raid():
    wb = Workbook()
    ws = sheet(wb, "Legend", first=True)
    legend(ws,
           ["1. One tab per RAID category: Risks, Assumptions, Issues, Dependencies.",
            "2. A risk becomes an issue only once it has materialised. Move the row, keep "
            "the originating risk ID.",
            "3. Severity is calculated as likelihood times impact. Do not type it.",
            f"4. Severity {C.ESCALATION_THRESHOLD} and above escalates to the Steering "
            f"Committee, per the integration charter.",
            "5. The Summary tab calculates automatically. Do not overwrite formulas."],
           [f"Reporting week {C.WEEK_LABEL}, day {C.DAYS_AFTER_DAY1} after Day 1 "
            f"({C.en(C.DAY1)}).",
            "Likelihood and impact scored 1 to 5. Bands: 1 to 6 Low, 8 to 12 Medium, "
            "15 to 20 High, 25 Critical.",
            "Date format DD.MM.YYYY throughout.",
            "This file is the single source of truth for risks. Where a status deck or a chat "
            "thread disagrees, the entry has to be corrected here first."])

    ws = sheet(wb, "Risks")
    titleblock(ws, "Risk Register",
               f"Status {C.en(C.TODAY)}  |  {C.R_TOTAL} open risks  |  "
               f"{C.R_HIGH} at or above severity {C.ESCALATION_THRESHOLD}  |  "
               f"Owner: {C.OFFICE} ({C.OFFICE_ABBR})")
    header(ws, 4, ["Risk ID", "Risk description", "Workstream", "Owner", "Likelihood",
                   "Impact", "Severity", "Band", "Mitigation action", "Due date",
                   "Status", "Trend", "Escalated to"],
           [9, 62, 15, 18, 12, 10, 11, 12, 58, 12, 13, 13, 18])
    lastr = 4 + len(C.RISKS)
    for i, r in enumerate(C.RISKS):
        row = 5 + i
        put(ws, row, [r[0], r[1], f"{r[3]} {C.WS_NAME[r[3]]}", C.nm(r[4]), r[5], r[6],
                      None, None, r[7], r[8], r[9], r[10], r[11] or ""], date_cols=(8,))
        c = ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX
        c = ws.cell(row=row, column=8,
                    value=f'=IF(G{row}>=25,"Critical",IF(G{row}>=15,"High",'
                          f'IF(G{row}>=8,"Medium","Low")))')
        c.fill = CALC
        c.font = RED if C.sev(r) >= 15 else AMB if C.sev(r) >= 8 else Font(A, 9)
        c.border = BOX

    ws = sheet(wb, "Assumptions")
    titleblock(ws, "Assumption Register",
               f"Assumptions the plan depends on and their validation status, "
               f"{C.en(C.TODAY)}")
    header(ws, 4, ["Assumption ID", "Assumption", "Workstream", "Date made",
                   "Impact if invalid", "Validation action", "Validate by",
                   "Validation status"],
           [14, 60, 14, 12, 60, 52, 13, 16])
    for i, a in enumerate(C.ASSUMPTIONS):
        put(ws, 5 + i, [a[0], a[1], f"{a[2]} {C.WS_NAME[a[2]]}", a[3], a[4], a[5], a[6], a[7]],
            date_cols=(4, 7))

    ws = sheet(wb, "Issues")
    titleblock(ws, "Issue Log",
               f"Risks that have materialised and problems already occurring, {C.en(C.TODAY)}")
    header(ws, 4, ["Issue ID", "Issue description", "Originating risk", "Date raised",
                   "Workstream", "Owner", "Severity", "Impact on milestone",
                   "Resolution action", "Target date", "Status", "Escalated to"],
           [9, 62, 14, 12, 14, 18, 11, 26, 54, 12, 13, 18])
    for i, x in enumerate(C.ISSUES):
        put(ws, 5 + i, [x[0], x[1], x[2] or "-", x[3], f"{x[4]} {C.WS_NAME[x[4]]}",
                        C.nm(x[5]), x[6], x[7], x[8], x[9], x[10], x[11] or ""],
            red_at=(7,), date_cols=(4, 10))

    ws = sheet(wb, "Dependencies")
    titleblock(ws, "Dependency Register",
               f"Cross-workstream handovers  |  decision B-03 requires both leads to confirm "
               f"|  {C.DEP_UNCONFIRMED_CRITICAL} still one-sided")
    header(ws, 4, ["Dependency ID", "Deliverable handed over", "From workstream",
                   "From owner", "To workstream", "To owner", "Needed by", "Criticality",
                   "Both leads confirmed", "Status"],
           [14, 62, 16, 18, 16, 18, 12, 13, 19, 13])
    for i, d in enumerate(C.DEPENDENCIES):
        put(ws, 5 + i, [d[0], d[1], f"{d[2]} {C.WS_NAME[d[2]]}", C.nm(d[2]),
                        f"{d[3]} {C.WS_NAME[d[3]]}", C.nm(d[3]), d[4], d[5], d[6], d[7]],
            red_at=(10,), date_cols=(7,))

    ws = sheet(wb, "Summary")
    titleblock(ws, "RAID Summary",
               "Calculates automatically from the four registers. No input.")
    ws.cell(row=4, column=1, value="Risks by severity band").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    header(ws, 5, ["Band", "Number", "Share"], [22, 14, 12])
    for i, b in enumerate(["Low", "Medium", "High", "Critical"]):
        r = 6 + i
        ws.cell(row=r, column=1, value=b).font = Font(A, 9)
        for col, f in ((2, f'=COUNTIF(Risks!$H$5:$H${lastr},$A{r})'),
                       (3, f'=IFERROR($B{r}/$B$10,"")')):
            c = ws.cell(row=r, column=col, value=f)
            c.fill = CALC
            c.font = Font(A, 9)
            c.border = BOX
            if col == 3:
                c.number_format = "0%"
    ws.cell(row=10, column=1, value="Total").font = Font(A, 9, bold=True)
    c = ws.cell(row=10, column=2, value="=SUM($B$6:$B$9)")
    c.fill = CALC
    c.font = Font(A, 9, bold=True)
    c.border = BOX

    ws.cell(row=12, column=1, value="RAID volume").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    for i, (lab, f) in enumerate([
        ("Open risks", f'=COUNTIF(Risks!$K$5:$K${lastr},"Open")'
                       f'+COUNTIF(Risks!$K$5:$K${lastr},"In progress")'),
        (f"Risks at or above severity {C.ESCALATION_THRESHOLD}",
         f'=COUNTIF(Risks!$G$5:$G${lastr},">={C.ESCALATION_THRESHOLD}")'),
        ("Risks escalated beyond the workstream",
         f'=COUNTIF(Risks!$M$5:$M${lastr},"Steering Committee")'
         f'+COUNTIF(Risks!$M$5:$M${lastr},"VCIO")'),
        ("Open assumptions", f'=COUNTIF(Assumptions!$H$5:$H${4+len(C.ASSUMPTIONS)},"Open")'),
        ("Open issues", f'=COUNTIF(Issues!$K$5:$K${4+len(C.ISSUES)},"Open")'
                        f'+COUNTIF(Issues!$K$5:$K${4+len(C.ISSUES)},"In progress")'),
        ("Dependencies at risk or delayed",
         f'=COUNTIF(Dependencies!$J$5:$J${4+len(C.DEPENDENCIES)},"At risk")'
         f'+COUNTIF(Dependencies!$J$5:$J${4+len(C.DEPENDENCIES)},"Delayed")'),
        ("Critical or high dependencies not confirmed by both leads",
         f'=COUNTIFS(Dependencies!$H$5:$H${4+len(C.DEPENDENCIES)},"Critical",'
         f'Dependencies!$I$5:$I${4+len(C.DEPENDENCIES)},"No")'
         f'+COUNTIFS(Dependencies!$H$5:$H${4+len(C.DEPENDENCIES)},"High",'
         f'Dependencies!$I$5:$I${4+len(C.DEPENDENCIES)},"No")'),
    ]):
        r = 13 + i
        ws.cell(row=r, column=1, value=lab).font = Font(A, 9)
        c = ws.cell(row=r, column=2, value=f)
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX
    ws.column_dimensions["A"].width = 52

    wb.save(C.OUT / "DellEMC_VCIO_RAID_Log_W3_2016-09-29.xlsx")


# =====================================================================
# 3. Synergy Tracker (English)
# =====================================================================
def synergy():
    wb = Workbook()
    ws = sheet(wb, "Legend", first=True)
    legend(ws,
           ["1. Baseline is locked and read only. It was signed off before Day 1.",
            "2. Synergy Register: one row per initiative. Target, secured and realised are "
            "entered per initiative, never at bucket level.",
            "3. Only Finance-validated initiatives count towards the reportable secured "
            "figure, per Steering Committee decision B-06.",
            "4. Phasing carries the in-year effect for the current financial year and "
            "cross-checks back to the register.",
            "5. Summary rolls up by bucket and reconciles to the deal model. No input."],
           [f"Reporting week {C.WEEK_LABEL}, day {C.DAYS_AFTER_DAY1} after Day 1.",
            "All amounts in USD million unless the column header says otherwise.",
            "Run-rate is the annualised effect once fully implemented. In-year is the effect "
            "landing in the current financial year.",
            f"Baseline source: synergy baseline sign-off dated {C.en(C.D_BASELINE)}, locked.",
            "Revenue synergies are designed at roughly three times cost synergies, consistent "
            "with the ratio management stated publicly at announcement. The absolute values "
            "are synthetic.",
            "Where this file and a status deck disagree, the Finance-validated figure here "
            "governs."])

    ws = sheet(wb, "Baseline")
    titleblock(ws, "Locked Financial Baseline",
               f"Baseline as at Day 1 ({C.en(C.DAY1)})  |  signed off "
               f"{C.en(C.D_BASELINE)}  |  do not edit")
    header(ws, 4, ["Baseline ID", "Line item", "Entity", "Cost category",
                   "FY16 actual (USD m)", "FY17 budget (USD m)", "Source document"],
           [13, 46, 20, 18, 18, 18, 42])
    for i, b in enumerate(C.BASELINE):
        put(ws, 5 + i, list(b))

    ws = sheet(wb, "Synergy Register")
    titleblock(ws, "Synergy Register",
               f"One row per initiative  |  {len(C.SYNERGIES)} initiatives  |  "
               f"status {C.en(C.D_SYNERGY)}  |  target USD {C.SYN_TARGET:,} m run-rate")
    header(ws, 4, ["Initiative ID", "Initiative", "Bucket", "Type", "Workstream", "Owner",
                   "Target run-rate (USD m)", "Secured run-rate (USD m)",
                   "Realised to date (USD m)", "Secured %", "FY17 in-year (USD m)",
                   "Cost to achieve (USD m)", "Status", "Finance validated"],
           [13, 56, 18, 10, 16, 18, 16, 17, 17, 11, 16, 17, 14, 15])
    lastsyn = 4 + len(C.SYNERGIES)
    for i, s in enumerate(C.SYNERGIES):
        r = 5 + i
        put(ws, r, [s[0], s[1], s[3], s[4], f"{s[5]} {C.WS_NAME[s[5]]}", C.nm(s[6]),
                    s[7], s[8], s[9], None, s[10], s[11], s[12], s[13]])
        c = ws.cell(row=r, column=10, value=f"=IFERROR(H{r}/G{r},\"\")")
        c.fill = CALC
        c.number_format = "0%"
        c.font = Font(A, 9)
        c.border = BOX
    r = lastsyn + 1
    ws.cell(row=r, column=2, value="Total").font = Font(A, 9, bold=True)
    for col in (7, 8, 9, 11, 12):
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({L}5:{L}{lastsyn})")
        c.fill = CALC
        c.font = Font(A, 9, bold=True)
        c.border = BOX
    c = ws.cell(row=r, column=10, value=f'=IFERROR(H{r}/G{r},"")')
    c.fill = CALC
    c.number_format = "0%"
    c.font = Font(A, 9, bold=True)
    c.border = BOX

    ws = sheet(wb, "Phasing")
    titleblock(ws, "In-Year Phasing, current financial year",
               "Monthly in-year effect per initiative in USD million, "
               "cross-checked against the register")
    months = ["Sep 16", "Oct 16", "Nov 16", "Dec 16", "Jan 17", "Feb 17"]
    header(ws, 4, ["Initiative ID", "Initiative"] + months + ["Total", "Check vs. register"],
           [13, 50] + [11] * 6 + [12, 18])
    for i, s in enumerate(C.SYNERGIES):
        r = 5 + i
        tot = s[10]
        # deterministic monthly split that always sums to the register value
        w = [0, 1, 2, 3, 4, 5]
        denom = sum(w)
        parts = [round(tot * x / denom) for x in w]
        parts[-1] += tot - sum(parts)
        put(ws, r, [s[0], s[1][:50]] + parts + [None, None])
        c = ws.cell(row=r, column=9, value=f"=SUM(C{r}:H{r})")
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX
        c = ws.cell(row=r, column=10,
                    value=f'=IF(I{r}=INDEX(\'Synergy Register\'!$K$5:$K${lastsyn},'
                          f'MATCH($A{r},\'Synergy Register\'!$A$5:$A${lastsyn},0)),'
                          f'"ok","check")')
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX
    r = lastsyn + 1
    ws.cell(row=r, column=2, value="Total in-year effect").font = Font(A, 9, bold=True)
    for col in range(3, 10):
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({L}5:{L}{lastsyn})")
        c.fill = CALC
        c.font = Font(A, 9, bold=True)
        c.border = BOX

    ws = sheet(wb, "Cost to Achieve")
    titleblock(ws, "Cost to Achieve",
               f"One-off cost required to deliver each initiative  |  total "
               f"USD {C.SYN_CTA:,} m")
    header(ws, 4, ["CTA ID", "Initiative ID", "Cost item", "Cost type", "Owner",
                   "Budgeted (USD m)", "Committed (USD m)", "Spent (USD m)",
                   "Remaining (USD m)"],
           [10, 13, 52, 20, 18, 15, 16, 14, 16])
    ctype = {"Procurement": "Internal effort", "IT cost": "External advisory",
             "Personnel": "Severance", "Footprint": "Site exit",
             "Operations": "Internal effort", "Revenue cross-sell": "Enablement",
             "Revenue pricing": "Enablement"}
    for i, s in enumerate(C.SYNERGIES):
        r = 5 + i
        budget = s[11]
        committed = round(budget * 0.7)
        spent = round(budget * 0.25)
        put(ws, r, [f"CTA-{i+1:02d}", s[0], f"Delivery cost for {s[1][:46]}",
                    ctype[s[3]], C.nm(s[6]), budget, committed, spent, None])
        c = ws.cell(row=r, column=9, value=f"=F{r}-H{r}")
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX
    r = lastsyn + 1
    ws.cell(row=r, column=3, value="Total").font = Font(A, 9, bold=True)
    for col in (6, 7, 8, 9):
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({L}5:{L}{lastsyn})")
        c.fill = CALC
        c.font = Font(A, 9, bold=True)
        c.border = BOX

    ws = sheet(wb, "Summary")
    titleblock(ws, "Synergy Summary by Bucket",
               "Calculates automatically from the register and the cost to achieve tab. "
               "No input.")
    header(ws, 5, ["Bucket", "Initiatives", "Target (USD m)", "Secured (USD m)", "Secured %",
                   "FY17 in-year (USD m)", "Cost to achieve (USD m)"],
           [24, 12, 16, 16, 12, 18, 20])
    for i, b in enumerate(C.SYN_BUCKETS):
        r = 6 + i
        ws.cell(row=r, column=1, value=b).font = Font(A, 9)
        for col, f in (
            (2, f"=COUNTIF('Synergy Register'!$C$5:$C${lastsyn},$A{r})"),
            (3, f"=SUMIF('Synergy Register'!$C$5:$C${lastsyn},$A{r},"
                f"'Synergy Register'!$G$5:$G${lastsyn})"),
            (4, f"=SUMIF('Synergy Register'!$C$5:$C${lastsyn},$A{r},"
                f"'Synergy Register'!$H$5:$H${lastsyn})"),
            (5, f'=IFERROR($D{r}/$C{r},"")'),
            (6, f"=SUMIF('Synergy Register'!$C$5:$C${lastsyn},$A{r},"
                f"'Synergy Register'!$K$5:$K${lastsyn})"),
            (7, f"=SUMIF('Synergy Register'!$C$5:$C${lastsyn},$A{r},"
                f"'Synergy Register'!$L$5:$L${lastsyn})"),
        ):
            c = ws.cell(row=r, column=col, value=f)
            c.fill = CALC
            c.font = Font(A, 9)
            c.border = BOX
            if col == 5:
                c.number_format = "0%"
    r = 6 + len(C.SYN_BUCKETS)
    ws.cell(row=r, column=1, value="Total").font = Font(A, 9, bold=True)
    for col in (2, 3, 4, 6, 7):
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({L}6:{L}{r-1})")
        c.fill = CALC
        c.font = Font(A, 9, bold=True)
        c.border = BOX
    c = ws.cell(row=r, column=5, value=f'=IFERROR($D{r}/$C{r},"")')
    c.fill = CALC
    c.number_format = "0%"
    c.font = Font(A, 9, bold=True)
    c.border = BOX

    base = r + 2
    ws.cell(row=base, column=1, value="Reconciliation to the deal model").font = \
        Font(A, 11, bold=True, color=C.D_DARK)
    lines = [
        ("Deal model target (USD m run-rate)", C.DEAL_MODEL_TARGET, None),
        ("Total target in register (USD m)", None, f"=$C${r}"),
        ("Gap to deal model (USD m)", None, f"=$B${base+1}-$B${base+2}"),
        ("Secured, all initiatives (USD m)", None, f"=$D${r}"),
        ("Secured, Finance validated only (USD m)", None,
         f"=SUMIF('Synergy Register'!$N$5:$N${lastsyn},\"Yes\","
         f"'Synergy Register'!$H$5:$H${lastsyn})"),
        ("Difference, not yet validated (USD m)", None, f"=$B${base+4}-$B${base+5}"),
        ("Cost synergy target (USD m)", None,
         f"=SUMIF('Synergy Register'!$D$5:$D${lastsyn},\"Cost\","
         f"'Synergy Register'!$G$5:$G${lastsyn})"),
        ("Revenue synergy target (USD m)", None,
         f"=SUMIF('Synergy Register'!$D$5:$D${lastsyn},\"Revenue\","
         f"'Synergy Register'!$G$5:$G${lastsyn})"),
        ("Revenue to cost ratio", None, f"=IFERROR($B${base+8}/$B${base+7},\"\")"),
    ]
    for i, (lab, val, f) in enumerate(lines):
        rr = base + 1 + i
        ws.cell(row=rr, column=1, value=lab).font = Font(A, 9)
        c = ws.cell(row=rr, column=2, value=val if val is not None else f)
        c.fill = CALC
        c.font = Font(A, 9)
        c.border = BOX
        if "ratio" in lab:
            c.number_format = "0.0"
    ws.cell(row=base + len(lines) + 2, column=1,
            value=f"Decision B-06: only the Finance-validated figure is reported to the "
                  f"Steering Committee from session 02 onward.").font = \
        Font(A, 8.5, italic=True, color=C.D_GREY)
    ws.column_dimensions["A"].width = 42

    wb.save(C.OUT / "DellEMC_VCIO_Synergy_Tracker_W3_2016-09-28.xlsx")


if __name__ == "__main__":
    tracker()
    raid()
    synergy()
    print("xlsx done")
