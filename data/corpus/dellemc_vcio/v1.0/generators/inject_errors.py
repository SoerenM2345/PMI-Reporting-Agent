"""Inject 10 deliberate flaws into a copy of the corpus and write the error key."""
import re
import shutil
import zipfile
from pathlib import Path

SRC = Path(__file__).resolve().parents[1] / "clean"
DST = Path(__file__).resolve().parents[1] / "with_errors"

LOG = []


def refresh():
    """Copy the clean corpus over, keeping the folder itself."""
    DST.mkdir(parents=True, exist_ok=True)
    for f in SRC.iterdir():
        if f.is_file():
            shutil.copy2(f, DST / f.name)


def patch_ooxml(fname, parts, pairs, eid):
    """String replacement inside the XML parts of a docx/pptx/xlsx."""
    p = DST / fname
    tmp = p.with_suffix(p.suffix + ".tmp")
    hits = {old: 0 for old, _ in pairs}
    with zipfile.ZipFile(p) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if any(re.match(pat, item.filename) for pat in parts):
                txt = data.decode("utf-8", "replace")
                for old, new in pairs:
                    if old in txt:
                        hits[old] += txt.count(old)
                        txt = txt.replace(old, new)
                data = txt.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(p)
    for old, n in hits.items():
        print(f"  [{eid}] {fname}: '{old[:44]}' replaced {n}x")
        assert n > 0, f"{eid}: pattern not found -> {old}"


def patch_text(fname, pairs, eid):
    p = DST / fname
    t = p.read_text(encoding="utf-8")
    for old, new in pairs:
        n = t.count(old)
        print(f"  [{eid}] {fname}: '{old[:44]}' replaced {n}x")
        assert n > 0, f"{eid}: pattern not found -> {old}"
        t = t.replace(old, new)
    p.write_text(t, encoding="utf-8")


DOCX = [r"word/document\.xml"]
PPTX = [r"ppt/slides/slide\d+\.xml"]
XLSX = [r"xl/sharedStrings\.xml", r"xl/worksheets/sheet\d+\.xml"]

WHR = "DellEMC_VCIO_Weekly_Highlight_Report_W3_2016-09-30.docx"
DECK = "DellEMC_VCIO_SteerCo_Update_Session02_2016-09-29.pptx"
RACI = "DellEMC_VCIO_RACI_Matrix_Integration_Hub_2016-09-23.html"
TRACKER = "DellEMC_VCIO_Integration_Tracker_W3_2016-09-29.xlsx"
PROT = "DellEMC_VCIO_Wochenprotokoll_KW39_2016-09-27.docx"
TERMS = "DellEMC_VCIO_Merger_Agreement_Key_Terms_2015-10-12.pdf"

refresh()
print("Injecting flaws\n")

# E-01 corrupted file -------------------------------------------------------
p = DST / TERMS
raw = p.read_bytes()
p.write_bytes(raw[: int(len(raw) * 0.55)])
print(f"  [E-01] {TERMS}: truncated {len(raw)} -> {len(raw)*55//100} bytes")

# E-02 file naming error ----------------------------------------------------
new_name = "DellEMC_VCIO_Wochenprotokoll_KW39_2016-10-27.docx"
(DST / PROT).rename(DST / new_name)
print(f"  [E-02] renamed {PROT} -> {new_name}")

# E-03 one unique position, two different people ----------------------------
patch_ooxml(WHR, DOCX, [("S. Lindqvist, VCIO Reporting Lead",
                         "P. Nakamura, VCIO Reporting Lead")], "E-03")

# E-04 Day 100 date wrong in the deck ---------------------------------------
patch_ooxml(DECK, PPTX, [("16 December 2016", "16 November 2016")], "E-04")

# E-05 flagship event date wrong in the weekly report ------------------------
patch_ooxml(WHR, DOCX, [("18 October 2016", "28 October 2016")], "E-05")

# E-06 Day 1 date wrong on the wiki page ------------------------------------
patch_text(RACI, [("7 September 2016", "9 September 2016")], "E-06")

# E-07 milestone count wrong in the deck ------------------------------------
patch_ooxml(DECK, PPTX, [("14 of 18", "15 of 18")], "E-07")

# E-08 synergy target digits transposed in the weekly report ----------------
patch_ooxml(WHR, DOCX, [("USD 7252 m", "USD 7522 m")], "E-08")

# E-09 task owner wrong in the tracker --------------------------------------
from openpyxl import load_workbook
wb = load_workbook(DST / TRACKER)
ws = wb["Massnahmenplan"]
done = False
for row in ws.iter_rows(min_row=5):
    if row[0].value == "A-022":
        assert row[3].value == "T. Bergström", row[3].value
        row[3].value = "K. Matsuda"
        done = True
assert done, "E-09: task A-022 not found"
wb.save(DST / TRACKER)
print("  [E-09] tracker: owner of A-022 changed T. Bergström -> K. Matsuda")

# E-10 risk severity wrong in the deck --------------------------------------
patch_ooxml(DECK, PPTX, [("4x5=20", "4x3=12")], "E-10")

print("\nAll 10 flaws injected.")
