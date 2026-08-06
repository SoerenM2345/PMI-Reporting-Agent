"""Write the error key workbook for the flawed corpus."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

DST = Path(__file__).resolve().parents[1]  # v1.0/ root — error key sits beside ground_truth.json
A = "Arial"
HDR = PatternFill("solid", fgColor="046A38")
ALT = PatternFill("solid", fgColor="F7F9F3")
PALE = PatternFill("solid", fgColor="F1F6E4")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROWS = [
    ("E-01", "Corrupted file", "Blocking",
     "DellEMC_VCIO_Merger_Agreement_Key_Terms_2015-10-12.pdf",
     "Whole file", "Valid 4-page PDF, 9,296 bytes",
     "Truncated to 5,112 bytes, no EOF marker, will not open",
     "The file fails to parse. The agent must report that it could not read it, not that "
     "it was empty",
     "Compare with the same file in ../clean"),

    ("E-02", "File naming error", "Medium",
     "DellEMC_VCIO_Wochenprotokoll_KW39_2016-10-27.docx",
     "Filename only, content untouched",
     "DellEMC_VCIO_Wochenprotokoll_KW39_2016-09-27.docx",
     "Date in the filename says 27.10.2016",
     "The filename says October, the document header and every date inside say 27.09.2016, "
     "and KW 39 is a September week",
     "Document metadata table inside the same file"),

    ("E-03", "One unique position held by two people", "High",
     "DellEMC_VCIO_Weekly_Highlight_Report_W3_2016-09-30.docx",
     "Metadata table, 'Prepared by'",
     "S. Lindqvist, VCIO Reporting Lead",
     "P. Nakamura, VCIO Reporting Lead",
     "There is exactly one VCIO Reporting Lead. Two different people are named in that role "
     "across the corpus. P. Nakamura is the Integration Risk Officer everywhere else",
     "RACI matrix section 4 (role holders), SteerCo transcript speaker list, SteerCo minutes "
     "attendance"),

    ("E-04", "Date error, reported vs. plan", "High",
     "DellEMC_VCIO_SteerCo_Update_Session02_2016-09-29.pptx",
     "Slide 2, executive summary key message",
     "16 December 2016 (Day 100)",
     "16 November 2016",
     "The deck states a Day 100 date one month earlier than the plan. Every other document "
     "carries 16 December 2016",
     "Integration roadmap (phases and gates), integration tracker (Meilensteine, M-17), "
     "weekly highlight report"),

    ("E-05", "Date error, reported vs. plan", "High",
     "DellEMC_VCIO_Weekly_Highlight_Report_W3_2016-09-30.docx",
     "Section 8, look ahead: milestone M-11 due date and the closing sentence",
     "18 October 2016 (flagship customer event, fixed external date)",
     "28 October 2016",
     "The weekly report moves a date that the roadmap marks as fixed and external. "
     "Two occurrences in the same document, both changed, so it reads as deliberate rather "
     "than a typo",
     "Integration roadmap (master milestone list, M-11), merger agreement key terms "
     "(key dates), integration tracker Meilensteine"),

    ("E-06", "Date error, reported vs. plan", "Medium",
     "DellEMC_VCIO_RACI_Matrix_Integration_Hub_2016-09-23.html",
     "Section 4, current role holders, 'Since' column (10 rows)",
     "7 September 2016 (Day 1)",
     "9 September 2016",
     "The wiki page dates every appointment to Day 1 but states Day 1 as 9 September. "
     "Day 1 is 7 September in the filing extract, the minutes, the tracker and the deck",
     "Merger agreement key terms (key dates), SteerCo minutes, integration tracker legend"),

    ("E-07", "Count inconsistent with the source of truth", "High",
     "DellEMC_VCIO_SteerCo_Update_Session02_2016-09-29.pptx",
     "Slide 2, executive summary status line",
     "Milestones on track 14 of 18",
     "Milestones on track 15 of 18",
     "The deck claims one more milestone on track than the tracker supports. Counting the "
     "milestone tab gives 14 (3 done plus 11 on track, against 3 delayed and 1 at risk)",
     "Integration tracker, Meilensteine tab and Auswertung tab; weekly highlight report "
     "section 1"),

    ("E-08", "Transposed digits in a headline figure", "High",
     "DellEMC_VCIO_Weekly_Highlight_Report_W3_2016-09-30.docx",
     "Section 1, overall assessment, 'Synergy secured' row",
     "USD 7252 m register target",
     "USD 7522 m",
     "Digits transposed. The synergy register sums to 7,252 and the same figure appears "
     "correctly in the deck and the tracker",
     "Synergy tracker, Summary tab total row; SteerCo deck value realisation slide"),

    ("E-09", "Owner assigned to the wrong workstream", "High",
     "DellEMC_VCIO_Integration_Tracker_W3_2016-09-29.xlsx",
     "Massnahmenplan tab, task A-022, column 'Verantwortlich'",
     "T. Bergström (Workstream Lead, IT)",
     "K. Matsuda (Workstream Lead, Supply Chain)",
     "Task A-022 sits in WS3 IT but is owned by the Supply Chain lead. The workstream column "
     "in the same row still says WS3 IT, so the row contradicts itself",
     "IT workstream one-pager, expert session record, escalation mail thread, RACI matrix"),

    ("E-10", "Risk severity inconsistent with the register", "High",
     "DellEMC_VCIO_SteerCo_Update_Session02_2016-09-29.pptx",
     "Slide 5, top risks table, risk R-01, column 'L x I'",
     "4x5=20, band High",
     "4x3=12, band shown as High but the arithmetic gives Medium",
     "The deck understates the top risk. The RAID log scores R-01 at likelihood 4, impact 5. "
     "The stated band no longer follows from the stated numbers, so the row is internally "
     "inconsistent as well",
     "RAID log, Risks tab; weekly highlight report section 5; SteerCo transcript"),
]

wb = Workbook()
ws = wb.active
ws.title = "Error key"
ws.sheet_view.showGridLines = False

ws["A1"] = "Dell-EMC synthetic PMI corpus - injected error key"
ws["A1"].font = Font(A, 14, bold=True, color="046A38")
ws["A2"] = ("10 deliberate flaws injected into a copy of the clean corpus. "
            "Reference corpus: ../clean. Everything not listed here is identical "
            "to the clean version.")
ws["A2"].font = Font(A, 9, color="75787B")

heads = ["ID", "Error type", "Severity", "File affected", "Where exactly",
         "Correct value", "Injected value", "How it is detectable",
         "Cross-check source in the corpus"]
widths = [7, 34, 11, 52, 44, 40, 44, 62, 52]
for j, h in enumerate(heads, 1):
    c = ws.cell(row=4, column=j, value=h)
    c.font = Font(A, 9, bold=True, color="FFFFFF")
    c.fill = HDR
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BOX
    ws.column_dimensions[get_column_letter(j)].width = widths[j - 1]
ws.row_dimensions[4].height = 30
ws.freeze_panes = "A5"

for i, row in enumerate(ROWS):
    r = 5 + i
    for j, v in enumerate(row, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(A, 9, bold=(j == 1))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BOX
        if r % 2 == 0:
            c.fill = ALT
        if j == 3 and v in ("High", "Blocking"):
            c.font = Font(A, 9, bold=True, color="DA291C")
    ws.row_dimensions[r].height = 58

# summary tab
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False
ws2["A1"] = "Summary"
ws2["A1"].font = Font(A, 13, bold=True, color="046A38")
summary = [
    ("Total flaws injected", len(ROWS)),
    ("Files touched", len({r[3] for r in ROWS})),
    ("Files in the folder", 21),
    ("Files identical to the clean corpus", 21 - len({r[3] for r in ROWS})),
    ("", ""),
    ("Corrupted / unreadable", sum(1 for r in ROWS if r[1].startswith("Corrupted"))),
    ("Naming errors", sum(1 for r in ROWS if "naming" in r[1])),
    ("Person and role conflicts", sum(1 for r in ROWS if "position" in r[1] or "Owner" in r[1])),
    ("Date errors, reported vs. plan", sum(1 for r in ROWS if r[1].startswith("Date error"))),
    ("Figure and count errors", sum(1 for r in ROWS if "Count" in r[1] or "Transposed" in r[1]
                                    or "severity" in r[1])),
    ("", ""),
    ("Detectable within a single document", 3),
    ("Detectable only by comparing two or more documents", 7),
]
for i, (k, v) in enumerate(summary):
    r = 3 + i
    ws2.cell(row=r, column=1, value=k).font = Font(A, 9, bold=(v == "" and k != ""))
    c = ws2.cell(row=r, column=2, value=v)
    c.font = Font(A, 9)
    if v != "":
        c.fill = PALE
        c.border = BOX
ws2.column_dimensions["A"].width = 52
ws2.column_dimensions["B"].width = 14
ws2.cell(row=18, column=1,
         value="Note: E-01 to E-10 are additional to the six conflicts that are planted in the "
               "clean corpus by design (see 00_README_Corpus.md section 5). Those six are not "
               "errors and are not listed here.").font = Font(A, 8.5, italic=True,
                                                              color="75787B")

wb.save(DST / "00_ERROR_KEY.xlsx")
print("error key written")
